import requests
import json
import os
import zipfile
import rarfile
import py7zr
import tarfile
from PyPDF2 import PdfReader
from docx import Document
from openpyxl import load_workbook
from PIL import Image

# 配置
APP_ID = "cli_a926e51e7178dcc0"
APP_SECRET = "nWEZIwPq8hayJJgjkgPScfnoBZdC6qDS"
CHAT_ID = "oc_2273a36513ca66dc7cbcebfe65135b21"
SAVE_DIR = "/root/.openclaw/workspace/attachments"
os.makedirs(SAVE_DIR, exist_ok=True)

def get_tenant_access_token():
    url = "https://open.feishu.cn/open-apis/auth/v3/tenant_access_token/internal"
    payload = json.dumps({"app_id": APP_ID, "app_secret": APP_SECRET})
    headers = {"Content-Type": "application/json"}
    response = requests.post(url, headers=headers, data=payload)
    return response.json()["tenant_access_token"]

def get_latest_attachments(token, limit=10):
    """获取最新的附件列表"""
    url = f"https://open.feishu.cn/open-apis/im/v1/messages?container_id_type=chat&container_id={CHAT_ID}&page_size={limit}"
    headers = {"Authorization": f"Bearer {token}"}
    response = requests.get(url, headers=headers)
    res = response.json()
    attachments = []
    for msg in res["data"]["items"]:
        if msg["msg_type"] == "file":
            content = json.loads(msg["body"]["content"])
            attachments.append({
                "message_id": msg["message_id"],
                "file_key": content["file_key"],
                "file_name": content["file_name"],
                "create_time": msg["create_time"]
            })
    return attachments

def download_attachment(token, message_id, file_key, file_name):
    """下载附件"""
    url = f"https://open.feishu.cn/open-apis/im/v1/messages/{message_id}/resources/{file_key}?type=file"
    headers = {"Authorization": f"Bearer {token}"}
    response = requests.get(url, headers=headers)
    save_path = os.path.join(SAVE_DIR, file_name)
    with open(save_path, "wb") as f:
        f.write(response.content)
    print(f"✅ 已下载: {save_path}")
    return save_path

def extract_archive(file_path):
    """自动解压压缩包"""
    ext = os.path.splitext(file_path)[1].lower()
    extract_dir = os.path.join(SAVE_DIR, os.path.splitext(os.path.basename(file_path))[0])
    os.makedirs(extract_dir, exist_ok=True)
    
    try:
        if ext == ".zip":
            with zipfile.ZipFile(file_path, 'r') as zf:
                zf.extractall(extract_dir)
        elif ext == ".rar":
            with rarfile.RarFile(file_path, 'r') as rf:
                rf.extractall(extract_dir)
        elif ext == ".7z":
            with py7zr.SevenZipFile(file_path, mode='r') as szf:
                szf.extractall(extract_dir)
        elif ext in [".tar", ".gz", ".bz2", ".xz"]:
            with tarfile.open(file_path, 'r:*') as tf:
                tf.extractall(extract_dir)
        else:
            return None
        
        print(f"✅ 已解压到: {extract_dir}")
        return extract_dir
    except Exception as e:
        print(f"❌ 解压失败: {e}")
        return None

def extract_text(file_path):
    """提取文档文本内容"""
    ext = os.path.splitext(file_path)[1].lower()
    text = ""
    
    try:
        if ext == ".txt" or ext == ".md":
            with open(file_path, 'r', encoding='utf-8') as f:
                text = f.read()
        elif ext == ".pdf":
            reader = PdfReader(file_path)
            for page in reader.pages:
                text += page.extract_text() + "\n"
        elif ext == ".docx":
            doc = Document(file_path)
            for para in doc.paragraphs:
                text += para.text + "\n"
        elif ext == ".xlsx":
            wb = load_workbook(file_path)
            for sheet in wb.sheetnames:
                ws = wb[sheet]
                text += f"=== Sheet: {sheet} ===\n"
                for row in ws.iter_rows(values_only=True):
                    text += "\t".join([str(cell) if cell else "" for cell in row]) + "\n"
        return text
    except Exception as e:
        print(f"❌ 文本提取失败: {e}")
        return None

def list_dir_tree(dir_path, prefix=""):
    """列出目录结构"""
    tree = ""
    items = sorted(os.listdir(dir_path))
    for i, item in enumerate(items):
        path = os.path.join(dir_path, item)
        is_last = i == len(items) - 1
        tree += f"{prefix}{'└── ' if is_last else '├── '}{item}\n"
        if os.path.isdir(path):
            tree += list_dir_tree(path, prefix + ('    ' if is_last else '│   '))
    return tree

def process_latest_attachment():
    """处理最新的附件"""
    token = get_tenant_access_token()
    attachments = get_latest_attachments(token, limit=5)
    
    if not attachments:
        print("❌ 未找到最近的附件")
        return
    
    # 处理最新的一个附件
    att = attachments[0]
    print(f"\n处理附件: {att['file_name']}")
    save_path = download_attachment(token, att["message_id"], att["file_key"], att["file_name"])
    
    # 尝试解压
    extract_dir = extract_archive(save_path)
    if extract_dir:
        print("\n📁 压缩包内容结构:")
        print(list_dir_tree(extract_dir))
    
    # 尝试提取文本
    text = extract_text(save_path)
    if text:
        print("\n📝 文档内容摘要:")
        print(text[:500] + "..." if len(text) > 500 else text)
    
    return save_path, extract_dir, text

if __name__ == "__main__":
    process_latest_attachment()
